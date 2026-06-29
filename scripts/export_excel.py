"""
export_excel.py — build a study-friendly vocab.xlsx from vocab/vocab.jsonl.

Layout (one sheet, frozen header + Word column, autofilter on for sort/filter):
  Word | Part of speech | IPA | Definition | Example | Known? | Topic | Date

Review tips this layout enables:
- Hide the Definition/Example columns to self-test from the Word.
- Filter "Known?" = (Blanks) to drill only the words you haven't learned yet.
- Tick "Known?" = Yes as you go (a dropdown keeps it consistent).

`--merge-from <file.xlsx>` carries over the "Known?" column (matched by word) from an
existing copy, so regenerating never wipes your progress.

Run:  python scripts/export_excel.py [--out vocab/vocab.xlsx] [--merge-from old.xlsx]
"""

import argparse
import json
import os

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
VOCAB_FILE = os.path.join(ROOT, "vocab", "vocab.jsonl")
DEFAULT_OUT = os.path.join(ROOT, "vocab", "vocab.xlsx")

# (header, jsonl key, width, wrap?)
COLUMNS = [
    ("Word", "word", 22, False),
    ("Part of speech", "pos", 16, False),
    ("IPA", "ipa", 22, False),
    ("Definition", "definition", 48, True),
    ("Example", "example", 52, True),
    ("Known?", "_known", 10, False),
    ("Topic", "topic", 34, True),
    ("Date", "date", 12, False),
]


def _load_rows():
    if not os.path.exists(VOCAB_FILE):
        return []
    rows = []
    with open(VOCAB_FILE, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                rows.append(json.loads(line))
            except json.JSONDecodeError:
                continue
    return rows


def _known_map(path):
    """Read {word_lower: known_value} from an existing export, if any."""
    if not path or not os.path.exists(path):
        return {}
    wb = load_workbook(path, read_only=True)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(max_row=1))]
    try:
        wi, ki = headers.index("Word"), headers.index("Known?")
    except ValueError:
        return {}
    out = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if wi < len(row) and ki < len(row) and row[wi] and row[ki]:
            out[str(row[wi]).strip().lower()] = row[ki]
    return out


def build(out_path, merge_from=None):
    rows = _load_rows()
    known = _known_map(merge_from)

    wb = Workbook()
    ws = wb.active
    ws.title = "Vocabulary"

    header_fill = PatternFill("solid", fgColor="1F2328")
    header_font = Font(bold=True, color="FFFFFF")
    wrap = Alignment(wrap_text=True, vertical="top")
    top = Alignment(vertical="top")

    # Header row
    for ci, (title, _key, width, _wrap) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=ci, value=title)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center")
        ws.column_dimensions[get_column_letter(ci)].width = width

    # Data rows
    for ri, r in enumerate(rows, start=2):
        for ci, (_title, key, _w, do_wrap) in enumerate(COLUMNS, start=1):
            if key == "_known":
                val = known.get((r.get("word") or "").strip().lower(), "")
            else:
                val = r.get(key, "")
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.alignment = wrap if do_wrap else top
            if key == "word":
                cell.font = Font(bold=True)

    last_row = len(rows) + 1
    ws.freeze_panes = "B2"  # keep header row + Word column visible while scrolling
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{last_row}"

    # "Known?" dropdown (column F)
    if last_row >= 2:
        dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"F2:F{last_row}")

    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    wb.save(out_path)
    return len(rows), out_path


def main():
    ap = argparse.ArgumentParser(description="Export vocab.jsonl to a study-friendly xlsx.")
    ap.add_argument("--out", default=DEFAULT_OUT, help="output .xlsx path")
    ap.add_argument("--merge-from", help="existing .xlsx to carry over the Known? column")
    args = ap.parse_args()

    n, path = build(args.out, args.merge_from)
    print(f"Wrote {n} words to {path}")


if __name__ == "__main__":
    main()
